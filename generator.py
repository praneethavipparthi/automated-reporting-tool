from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os

# Function to calculate column index from the column letter
def column_index(col_letter):
    return ord(col_letter) - ord('A') + 1

# Function to generate a formula for the sum of a column
def sum_formula(column_letter, start_row, end_row):
    return f'=SUM({column_letter}{start_row}:{column_letter}{end_row})'

def get_user_input(prompt, data_type=str):
    while True:
        try:
            user_input = data_type(input(prompt))
            return user_input
        except ValueError:
            print("Invalid input. Please enter a valid value.")

# Input month and categories
month = get_user_input("Enter the month: ")

# Input categories and sales data
categories_str = input("Enter categories (comma-separated): ")
categories = categories_str.split(',')

data = []
for category in categories:
    sales = get_user_input(f"Enter sales for {category}: ", int)
    data.append([sales for _ in range(len(categories))])

# Create a new Workbook
wb = Workbook()
sheet = wb.active

# Input data into the sheet
for row_index, row_data in enumerate(data, start=2):
    for col_index, value in enumerate(row_data, start=2):
        sheet.cell(row=row_index, column=col_index, value=value)

# Create a bar chart
min_row = 1
max_row = len(data) + 1
min_column = column_index('B')  # Assuming data starts from column B
max_column = min_column + len(categories) - 1

barchart = BarChart()
data_ref = Reference(sheet, min_col=min_column, min_row=min_row, max_col=max_column, max_row=max_row)
categories_ref = Reference(sheet, min_col=min_column, min_row=min_row + 1, max_row=max_row)
barchart.add_data(data_ref, titles_from_data=True)
barchart.set_categories(categories_ref)

# Add the chart to the sheet
sheet.add_chart(barchart, "B12")
barchart.title = f'Sales by Product line - {month}'
barchart.style = 5  # Choose the chart style

# Write sum formulas with a for loop
for col_index in range(min_column + 1, max_column + 1):
    col_letter = get_column_letter(col_index)
    sheet[f'{col_letter}{max_row + 1}'] = sum_formula(col_letter, min_row + 1, max_row)
    sheet[f'{col_letter}{max_row + 1}'].style = 'Currency'

# Add formatting
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

# Save the workbook
application_path = os.path.dirname(os.path.abspath(__file__))  # Adjust as needed
output_path = os.path.join(application_path, f'report_{month}.xlsx')
wb.save(output_path)
